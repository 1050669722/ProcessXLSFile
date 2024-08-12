import os, sys

from openpyxl import load_workbook
from openpyxl.styles import Font
from typing import List


def get_rows_with_red_font_color(xls_path, sheet_name, outputFilePath):
    # 加载工作簿和指定的工作表
    wb = load_workbook(xls_path)
    ws = wb[sheet_name]

    # 存储包含红色字体的行号
    red_font_rows = []

    # 遍历工作表中的所有单元格
    for row in ws.iter_rows():
        for cell in row:
            # 检查单元格的字体颜色是否为红色
            if cell.font and cell.font.color and cell.font.color.rgb == 'FFFF0000':
                red_font_rows.append(ws.cell(row=row[0].row, column=1).row)  # 假设检查第一列的单元格颜色
                break  # 找到红色字体后，跳出当前行的循环

    # 搜集每一行的字符串
    wordPairs = []
    for rowNumber in red_font_rows:
        getStringAccordingToRowNumber(ws, rowNumber, wordPairs)
    # print(wordPairs)

    # 将这些字符串写入文件
    with open(os.path.join(outputFilePath, "已经标注为红色的六选二词对.txt"), 'w', encoding='utf-8') as file:
        for wordPair in wordPairs:
            file.write(wordPair)
            file.write("\n")

    return red_font_rows


def getStringAccordingToRowNumber(workSheet, rowNumber, wordPairs: List[str]):
    # 检查行号是否有效
    if workSheet.max_row >= rowNumber:
        # 获取指定行
        row = workSheet[rowNumber]

        # 遍历行中的所有单元格
        currentLine = ""
        for cell in row:
            # 获取单元格中的字符串
            # print(cell.value)
            if cell.value:
                currentLine += cell.value
                currentLine += ", "
        currentLine += "\n"
        wordPairs.append(currentLine)

    else:
        print(f"Row number {rowNumber} is out of range.")


if __name__ == "__main__":
    # 使用示例
    xls_path = R"D:\files\GRE\语文\六选二词对.xlsx"
    sheet_name = '潘晨光GRE终极词表'
    outputFilePath = os.path.dirname(xls_path)
    red_rows = get_rows_with_red_font_color(xls_path, sheet_name, outputFilePath)
    print("Rows with red font color:", red_rows)
    print("The length of red_rows is: {}".format(len(red_rows)))

    # 程序结束前的提示
    print("按任意键退出程序...")
    # 使用sys.stdin.read()或input()来等待按键
    # sys.stdin.read(1)会读取一个字符，但不会显示在控制台上
    # 如果你想要看到用户按下的字符，可以使用input()，但会显示在控制台上
    input_char = sys.stdin.read(1)  # 或者使用 input()
    # 这里可以添加一些清理代码，如果需要的话
    # 程序退出
    sys.exit()
